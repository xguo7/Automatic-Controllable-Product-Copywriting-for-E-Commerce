import openpyxl
import random
import re
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.worksheet.datavalidation import DataValidation

all_data=[]
for i in range(9):
        with open('/home/xiaojie.guo/daren_data/assigned_write_1343/assigned_write'+str(i)+'.txt', encoding='utf-8') as f:
            data=f.readlines()
            random.shuffle(data)   
            all_data.extend(data[:2000])


random.shuffle(all_data)

def write_excel(filename,data):
        workbook=openpyxl.Workbook()
        worksheet=workbook.active

        title=['类别','文案','商品信息','sku']
        for i in range(len(title)):
            worksheet.cell(1,i+1,title[i])

        total_c=0
        for item in data:
            item=item.split('|||')
            prod=item[1]
            prod=ILLEGAL_CHARACTERS_RE.sub(r'', prod)
            write=item[2]   
            sku=item[0]
            #worksheet.cell(total_c+2,1,'0机身外观')
            worksheet.cell(total_c+2,2,write)
            worksheet.cell(total_c+2,3,prod)
            worksheet.cell(total_c+2,4,sku)
            total_c+=1

        #dv=DataValidation(type='list', formula1='"0机身外观","1屏幕音效","2网络双模5G","3照相美颜","4性能","5电池","6解锁"',allow_blank=True)
        #worksheet.add_data_validation(dv) 
        #dv.add('A2:A301')

        workbook.save(filename)


for i in range(61):
    write_excel('/home/xiaojie.guo/daren_data/assigned_write_1343/human_label_1343/write_test'+str(i)+'.xlsx',all_data[i*300:i*300+300])
